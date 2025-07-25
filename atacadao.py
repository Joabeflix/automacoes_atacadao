import pandas as pd
from openpyxl import load_workbook

df = pd.DataFrame(columns=["VENCIMENTO", "CTRL HISTORICO", "FORMA PAGTO", "ENTRADA"])

arquivo = "relatorio_simples_com_cabecalho.xlsx"
df.to_excel(arquivo, sheet_name="Relatório", startrow=5, index=False)

wb = load_workbook(arquivo)
ws = wb["Relatório"]

ws["A1"] = "ATACADAO AUTOMOTIVA LTDA"
ws["A2"] = "CNPJ 49.973.821/0001-01"
ws["D2"] = "BANCO: 748 AG: 0167 CC: 80116-1"

ws["A3"] = "CNPJ 49.973.821/0002-84"
ws["D3"] = "BANCO: 748 AG: 0167 CC: 40358-3"

ws["A4"] = "CNPJ 49.973.821/0003-65"
ws["D4"] = "BANCO: 748 AG: 0167 CC: 55493-1"

ws["A5"] = "PERIODO: 01/01/2024 ATÉ 31/08/2024"
ws["D5"] = "CARTÃO DE CRÉDITO"

wb.save(arquivo)
